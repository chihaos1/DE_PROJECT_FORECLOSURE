from openpyxl import Workbook
from django.shortcuts import render
from django.http import HttpResponse
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework.response import Response
from .serializer import EquatorPropertiesSerializer
from .models import EquatorProperties, RealtorProperties
from datetime import date
from foreclosure.foreclosure_src.main import main


class EquatorPropertiesView(APIView):
    def get(self, request):
        output = [
            {"street_name": output.street_name}
            for output in EquatorProperties.objects.all()
        ]
        return Response(output)

    def post(self, request):
        serializer = EquatorPropertiesSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response(serializer)


@api_view(["GET"])
def get_foreclosure(request):
    equ_output = [
        {
            "source": output.source,
            "property_url": output.property_url,
            "street_name": output.street_name,
            "city": output.city,
            "state": output.state,
            "zip": output.zip,
            "listed_price": output.listed_price,
            "year_built": output.year_built,
            "property_type": output.property_type,
            "sqft": output.square_ft,
            "bedroom": output.bedroom,
            "bathroom": output.bathroom,
            "photo_one_url": output.photo_one,
        }
        for output in EquatorProperties.objects.all()
    ]

    rea_output = [
        {
            "source": output.source,
            "property_url": output.property_url,
            "street_name": output.street_name,
            "city": output.city,
            "state": output.state,
            "zip": output.zip,
            "listed_price": output.listed_price,
            "year_built": output.year_built,
            "property_type": output.property_type,
            "sqft": output.square_ft,
            "bedroom": output.bedroom,
            "bathroom": output.bathroom,
            "photo_one_url": output.photo_one,
        }
        for output in RealtorProperties.objects.all()
    ]

    output = equ_output + rea_output

    return Response(output)


@api_view(["GET"])
def scrape(request, location):
    main(location)
    response = HttpResponse(
        content_type="application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f"attachment; filename=PropertyList{date.today().strftime('%Y%m%d')}.xlsx"
    )
    equ_property_queryset = EquatorProperties.objects.all().values()
    rea_property_queryset = RealtorProperties.objects.all().values()

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet(index=1, title="Equator")
    worksheet = workbook["Equator"]
    equ_columns = [
        "Source",
        "Property URL",
        "Street Name",
        "City",
        "State",
        "Zip",
        "County",
        "Listed Price",
        "Year Built",
        "Property Type",
        "Description",
        "Bedroom",
        "Bathroom",
        "Square Feet",
        "Lot Size Square Feet",
        "Lot Size Acre",
        "Has Basement",
        "Has AC",
        "Has Pool",
        "Estimated Cashflow",
        "Estimated Gross Yield",
        "Estimated Net Yield",
        "MLS Number",
        "Realty",
        "Agent Name",
        "Agent Email",
        "Agent Phone",
        "Property Image 1",
        "Property Image 2",
        "Property Image 3"
    ]
    worksheet.append(equ_columns)

    for property in equ_property_queryset:
        property_details = list(property.values())
        worksheet.append(property_details[1:-1]) #except ID and search date

    workbook.create_sheet(index=2, title="Realtor")
    worksheet = workbook["Realtor"]
    rea_columns = [
        "Source",
        "Property URL",
        "Street Name",
        "City",
        "State",
        "Zip",
        "County",
        "Listed Price",
        "Year Built",
        "Property Type",
        "Description",
        "Bedroom",
        "Bathroom",
        "Square Feet",
        "Stories",
        "Garage",
        "Lot Size Square Feet",
        "Monthly Payment",
        "Principal & Interest",
        "Home Insurance",
        "HOA Fees",
        "Property Tax",
        "Agent Name",
        "Agent Email",
        "Property Image"
    ]
    worksheet.append(rea_columns)

    for property in rea_property_queryset:
        property_details = list(property.values())
        worksheet.append(property_details[1:-1])

    workbook.save(response)
    return response
